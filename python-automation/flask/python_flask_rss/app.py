from flask import Flask, render_template, request
from flask import send_file
import googletrans
import openpyxl
import os
import time
import feedparser

app = Flask(__name__) #초기화

@app.route('/') #라우터
def index():
    return render_template('index.html')

@app.route('/rss', methods=['GET', 'POST']) #http://127.0.0.1:5000/rss
def rss():
    rss_url = request.form['rss_url'] #request(index.html의 form data)에서 name=rss_url 인 input값을 rss_url로 지정
    feed = feedparser.parse(rss_url) 
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = "제목"
    worksheet['B1'] = "링크"
    worksheet['C1'] = "내용"

    for row in range(2, 20):
        entry = feed.entries[row]
        worksheet.cell(row=row, column=1, value=entry.title)
        worksheet.cell(row=row, column=2, value=entry.link)
        worksheet.cell(row=row, column=3, value=entry.description)
    
    workbook.save("rss_result.xlsx")
    return process("rss_result.xlsx")

@app.route('/process', methods=['GET', 'POST'])
def process(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active 
    translator = googletrans.Translator()
    for row in sheet.iter_rows(): #엑셀에서 데이터가 존재하는 곳까지 for문 돌기
        for cell in row:
            translated_text = translator.translate(cell.value, dest='en').text #cell 값을 읽어 영어로 바꾼 후, text를 가져옴
            cell.value = translated_text #번역된 text로 cell 값을 update
        time.sleep(0.5)

    workbook.save('transrated_rss.xlsx')

    return render_template("result.html", file = file) #result.html로 파일들을 보내

@app.route('/download_report')
def download_report():
    return send_file('transrated_rss.xlsx', as_attachment=True) #as_attachemnt: 첨부파일 형식으로 보내겠다

if __name__ == "__main__":
    app.run(debug=True) #디버깅 용으로 동작